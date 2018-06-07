#! python3
#-*- coding: gbk -*-

from urllib.request import urlretrieve
from urllib.request import urlopen
from bs4 import BeautifulSoup

import re
import os
import openpyxl

#   确认文件夹存在，不存在，就创建它们
teacherfolder = "D:\结构所教师"
teacherfolder0 = "D:\结构所教师\教授"
teacherfolder1 = "D:\结构所教师\副教授"
teacherfolder2 = "D:\结构所教师\讲师"
teacherfolder3 = "D:\结构所教师\兼职教授"

if not os.path.exists(teacherfolder):
    try:
        os.makedirs(teacherfolder)
    except OSError:
        pass
        
if not os.path.exists(teacherfolder0):
    try:
        os.makedirs(teacherfolder0)
    except OSError:
        pass

if not os.path.exists(teacherfolder1):
    try:
        os.makedirs(teacherfolder1)
    except OSError:
        pass

if not os.path.exists(teacherfolder2):
    try:
        os.makedirs(teacherfolder2)
    except OSError:
        pass

if not os.path.exists(teacherfolder3):
    try:
        os.makedirs(teacherfolder3)
    except OSError:
        pass

imagelocation0 = []
imagelocation1 = []
imagelocation2 = []
imagelocation3 = []

teachernames0 = []
teachernames1 = []
teachernames2 = []
teachernames3 = []

teacherhomepage0 = []
teacherhomepage1 = []
teacherhomepage2 = []
teacherhomepage3 = []

#   处理副教授信息
#   info_lb=
#   50教授，51副教授，52讲师，53兼职教授

#   教授==========
url = 'http://risedr.tongji.edu.cn/pictrue.aspx?info_lb=50&flag=3'
html = urlopen(url)
bsObj = BeautifulSoup(html, "html.parser")

#   获取老师照片地址
images = bsObj.findAll("img",{"src":re.compile("uploadfiles\/.*\.jpg")})
for image in images:
    imagelocation0.append('http://risedr.tongji.edu.cn/' + image["src"])
    
#   获取老师姓名地址 和 个人主页
teachernamelist= bsObj.findAll("div",{"class":"title"})
for teachername in teachernamelist :
    teachernames0.append(teachername.a.get_text().strip())
    teacherhomepage0.append('http://risedr.tongji.edu.cn/' + teachername.a["href"])

#   保存老师照片到文件夹
for i in range(len(imagelocation0)):
    urlretrieve(imagelocation0[i], teacherfolder0 + "\\" + teachernames0[i] + ".jpg")
    
teachernumber0 = len(teachernames0)


#   副教授==========
url = 'http://risedr.tongji.edu.cn/pictrue.aspx?info_lb=51&flag=3'
html = urlopen(url)
bsObj = BeautifulSoup(html, "html.parser")

#   获取老师照片地址
images = bsObj.findAll("img",{"src":re.compile("uploadfiles\/.*\.jpg")})
for image in images:
    imagelocation1.append('http://risedr.tongji.edu.cn/' + image["src"])
    
    
#   获取老师姓名地址 和 个人主页
teachernamelist= bsObj.findAll("div",{"class":"title"})
for teachername in teachernamelist :
    teachernames1.append(teachername.a.get_text().strip())
    teacherhomepage1.append('http://risedr.tongji.edu.cn/' + teachername.a["href"])
    
#   保存老师照片到文件夹
for i in range(len(imagelocation1)):
    urlretrieve(imagelocation1[i], teacherfolder1 + "\\" + teachernames1[i] + ".jpg")
    
teachernumber1 = len(teachernames1)

#   讲师==========
url = 'http://risedr.tongji.edu.cn/pictrue.aspx?info_lb=52&flag=3'
html = urlopen(url)
bsObj = BeautifulSoup(html, "html.parser")

#   获取老师照片地址
images = bsObj.findAll("img",{"src":re.compile("uploadfiles\/.*\.jpg")})
for image in images:
    imagelocation2.append('http://risedr.tongji.edu.cn/' + image["src"])
    
    
#   获取老师姓名地址 和 个人主页
teachernamelist= bsObj.findAll("div",{"class":"title"})
for teachername in teachernamelist :
    teachernames2.append(teachername.a.get_text().strip())
    teacherhomepage2.append('http://risedr.tongji.edu.cn/' + teachername.a["href"])
    
#   保存老师照片到文件夹
for i in range(len(imagelocation2)):
    urlretrieve(imagelocation2[i], teacherfolder2 + "\\" + teachernames2[i] + ".jpg")
    
teachernumber2 = len(teachernames2)

#   兼职教授==========
url = 'http://risedr.tongji.edu.cn/pictrue.aspx?info_lb=53&flag=3'
html = urlopen(url)
bsObj = BeautifulSoup(html, "html.parser")

#   获取老师照片地址
images = bsObj.findAll("img",{"src":re.compile("uploadfiles\/.*\.jpg")})
for image in images:
    imagelocation3.append('http://risedr.tongji.edu.cn/' + image["src"])
    
#   获取老师姓名地址 和 个人主页
teachernamelist= bsObj.findAll("div",{"class":"title"})
for teachername in teachernamelist :
    teachernames3.append(teachername.a.get_text().strip())
    teacherhomepage3.append('http://risedr.tongji.edu.cn/' + teachername.a["href"])
    
#   保存老师照片到文件夹
for i in range(len(imagelocation3)):
    urlretrieve(imagelocation3[i], teacherfolder3 + "\\" + teachernames3[i] + ".jpg")
    
teachernumber3 = len(teachernames3)

    
#   生成excel表格
wb = openpyxl.Workbook()

teachersheet0 = wb.create_sheet(index = 0, title = "教授信息")
teachersheet1 = wb.create_sheet(index = 1, title = "副教授信息")
teachersheet2 = wb.create_sheet(index = 2, title = "讲师信息")
teachersheet3 = wb.create_sheet(index = 3, title = "兼职教授信息")

teachersheet0['A1'] = '教师姓名'
teachersheet0['B1'] = '教师个人主页'
for row in range(2, teachernumber0 + 2):
    teachersheet0.cell(column = 1, row = row, value = teachernames0[row - 2])
    teachersheet0.cell(column = 2, row = row, value = teacherhomepage0[row - 2])
teachersheet0.column_dimensions['A'].width = 18
teachersheet0.column_dimensions['B'].width = 75

teachersheet1['A1'] = '教师姓名'
teachersheet1['B1'] = '教师个人主页'
for row in range(2, teachernumber1 + 2):
    teachersheet1.cell(column = 1, row = row, value = teachernames1[row - 2])
    teachersheet1.cell(column = 2, row = row, value = teacherhomepage1[row - 2])
teachersheet1.column_dimensions['B'].width = 75

teachersheet2['A1'] = '教师姓名'
teachersheet2['B1'] = '教师个人主页'
for row in range(2, teachernumber2 + 2):
    teachersheet2.cell(column = 1, row = row, value = teachernames2[row - 2])
    teachersheet2.cell(column = 2, row = row, value = teacherhomepage2[row - 2])
teachersheet2.column_dimensions['A'].width = 26
teachersheet2.column_dimensions['B'].width = 75

teachersheet3['A1'] = '教师姓名'
teachersheet3['B1'] = '教师个人主页'
for row in range(2, teachernumber3 + 2):
    teachersheet3.cell(column = 1, row = row, value = teachernames3[row - 2])
    teachersheet3.cell(column = 2, row = row, value = teacherhomepage3[row - 2])
teachersheet3.column_dimensions['A'].width = 28
teachersheet3.column_dimensions['B'].width = 75

wb.save(teacherfolder + '\结构所老师个人信息.xlsx')


